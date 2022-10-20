import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook

#getting input from user for search
userinput = input('please input search:- ')

browser = webdriver.Chrome('chromedriver.exe')
browser.maximize_window()
browser.get("https://www.google.com/en")

browser.find_element(By.NAME, 'q').send_keys(userinput)

time.sleep(3)

for element in browser.find_elements(By.XPATH, '//div[@class="mkHrUc"]'):
    title = browser.find_element(By.XPATH, './/ul').text
    # print(title)
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells

    ws['A1'] = datetime.date.today()

    ws['A2'] = title

    ws['B1'] = userinput
    # Save the file
    wb.save("Excel.xlsx")

print("Excel sheet generate successfully for :- "+userinput)
browser.close()
